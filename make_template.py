"""Build sports-daily-control.xlsx, the starting point for the Google Sheet.

Upload it to Drive and open it with Google Sheets (or import it into a new
sheet), then paste the spreadsheet id into config.json. Re-run this only if you
want a fresh blank copy -- it overwrites the file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "sports-daily-control.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(color="808080", italic=True)

SPORTS = "NFL,NBA,MLB,NHL,EPL,MLS,USMNT,CFB,CBB,CHKY"
TIERS = "favorite,follow,watch"
OPTIONS = ("enabled,show_all_games,national_tv_only,standalone_only,"
           "include_postseason,playoff_race,race_from_month,race_last_days,"
           "race_only_last_spot,race_until_settled,race_until_season_end,race_min_odds,"
           "hide_finished,show_odds,show_records,timezone")


def _header(ws, labels, widths):
    ws.append(labels)
    for idx, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"


def teams_tab(wb):
    ws = wb.active
    ws.title = "Teams"
    _header(ws, ["Sport", "Team", "Tier", "Note", "Expires"], [10, 30, 12, 34, 14])

    rows = [
        ["NFL", "Detroit Lions", "favorite", "", ""],
        ["MLB", "Detroit Tigers", "favorite", "", ""],
        ["NBA", "Detroit Pistons", "favorite", "", ""],
        ["NBA", "Cleveland Cavaliers", "follow", "", ""],
        ["NHL", "Detroit Red Wings", "favorite", "", ""],
        ["MLB", "", "favorite", "", ""],
        ["NBA", "", "favorite", "", ""],
        ["NHL", "", "favorite", "", ""],
        ["EPL", "Tottenham Hotspur", "favorite", "", ""],
        ["MLS", "Atlanta United", "favorite", "", ""],
        ["USMNT", "United States", "favorite", "", ""],
        ["CFB", "Michigan Wolverines", "favorite", "", ""],
        ["CFB", "Cornell Big Red", "favorite", "", ""],
        ["CFB", "Ohio State Buckeyes", "watch", "Rival", ""],
        ["CFB", "Michigan State Spartans", "watch", "Rival", ""],
        ["CFB", "Notre Dame Fighting Irish", "watch", "Rival", ""],
        ["CFB", "Syracuse Orange", "watch", "Other", ""],
        ["CBB", "Michigan Wolverines", "favorite", "", ""],
        ["CBB", "Cornell Big Red", "favorite", "", ""],
        ["CBB", "Ohio State Buckeyes", "watch", "Rival", ""],
        ["CBB", "Michigan State Spartans", "watch", "Rival", ""],
        ["CBB", "Notre Dame Fighting Irish", "watch", "Rival", ""],
        ["CBB", "Syracuse Orange", "watch", "Other", ""],
        ["CHKY", "Michigan Wolverines", "favorite", "", ""],
        ["CHKY", "Cornell Big Red", "favorite", "", ""],
    ]
    for row in rows:
        ws.append(row)

    sport_dv = DataValidation(type="list", formula1='"%s"' % SPORTS, allow_blank=True)
    tier_dv = DataValidation(type="list", formula1='"%s"' % TIERS, allow_blank=True)
    ws.add_data_validation(sport_dv)
    ws.add_data_validation(tier_dv)
    sport_dv.add("A2:A200")
    tier_dv.add("C2:C200")

    ws.append([])
    ws.append(["", "Sample rows -- replace them with your own.", "", "", ""])
    ws.append(["", "Tier 'favorite' pins the game at the top; 'watch' puts it "
                   "in Key opponents with the Note shown as a badge.", "", "", ""])
    ws.append(["", "Expires is optional (YYYY-MM-DD). A past date drops the row "
                   "automatically, so playoff-race entries clean themselves up.", "", "", ""])
    for offset in range(3):
        ws.cell(row=ws.max_row - offset, column=2).font = NOTE_FONT
    return ws


def options_tab(wb):
    ws = wb.create_sheet("Options")
    _header(ws, ["Scope", "Option", "Value", "Notes"], [10, 22, 18, 44])

    rows = [
        ["All", "timezone", "America/New_York", "IANA name"],
        ["All", "show_records", "TRUE", "W-L next to each team"],
        ["All", "show_odds", "TRUE", "spread and over/under"],
        ["All", "hide_finished", "FALSE", "drop games already over (favorites always stay)"],
        ["NFL", "show_all_games", "FALSE", "not the regional Sunday windows"],
        ["NFL", "standalone_only", "TRUE", "TNF, SNF, MNF, Thanksgiving, Sat specials"],
        ["NFL", "include_postseason", "TRUE", "every playoff game"],
        ["NFL", "playoff_race", "TRUE", "adds who the Lions are competing with"],
        ["NFL", "race_from_month", "11", "November onward only"],
        ["NFL", "race_min_odds", "20", "only while the Lions are 20%+ to make it"],
        ["NBA", "show_all_games", "FALSE", "no neutral games"],
        ["NBA", "playoff_race", "TRUE", ""],
        ["NBA", "race_from_month", "3", "March onward"],
        ["NBA", "race_until_season_end", "TRUE", "until the regular season actually ends"],
        ["NBA", "race_only_last_spot", "TRUE", "just the team holding the last spot"],
        ["NBA", "race_until_settled", "TRUE", "until the Pistons clinch or are out"],
        ["NHL", "show_all_games", "FALSE", "no neutral games"],
        ["NHL", "playoff_race", "TRUE", ""],
        ["NHL", "race_from_month", "3", "March onward"],
        ["NHL", "race_until_season_end", "TRUE", "until the regular season actually ends"],
        ["NHL", "race_only_last_spot", "TRUE", ""],
        ["NHL", "race_until_settled", "TRUE", ""],
        ["EPL", "show_all_games", "TRUE", ""],
        ["MLB", "show_all_games", "FALSE", "15 a night is too many"],
        ["CFB", "include_postseason", "TRUE", "all bowls and CFP games"],
        ["CBB", "include_postseason", "FALSE", "March Madness uses the regular rules"],
        ["CHKY", "include_postseason", "TRUE", "NCAA regionals and Frozen Four"],
        ["MLB", "playoff_race", "TRUE", "AL Central and the wild card"],
        ["MLB", "race_from_month", "7", "July onward only"],
        ["MLB", "race_min_odds", "20", "only while the Tigers are 20%+ to make it"],
    ]
    for row in rows:
        ws.append(row)

    scope_dv = DataValidation(type="list", formula1='"All,%s"' % SPORTS, allow_blank=True)
    option_dv = DataValidation(type="list", formula1='"%s"' % OPTIONS, allow_blank=True)
    ws.add_data_validation(scope_dv)
    ws.add_data_validation(option_dv)
    scope_dv.add("A2:A100")
    option_dv.add("B2:B100")
    return ws


def readme_tab(wb):
    ws = wb.create_sheet("How it works")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 88

    lines = [
        ("Teams tab", "Sport, Team, Tier, Note, Expires. One row per team."),
        ("", "Tier 'favorite' = your teams, pinned at the top of the page."),
        ("", "Tier 'follow' = pinned like a favorite, but never drives the playoff"),
        ("", "race -- for a team whose games you want without their race."),
        ("", "Tier 'watch' = key opponents; the Note becomes the badge explaining why."),
        ("", "Usually you can leave 'watch' empty -- playoff_race fills Key opponents"),
        ("", "on its own. Add a row only for something the standings cannot know."),
        ("", "Expires (optional, YYYY-MM-DD) drops the row once the date passes."),
        ("", "Names are checked against ESPN -- run sports_daily.py --check after editing."),
        ("", "Use FULL names for college: matching is substring, so 'Michigan' also"),
        ("", "means Michigan State and six others. --check flags that."),
        ("", "EPL / MLS / USMNT entries cover every competition that team plays in --"),
        ("", "domestic cups, Europe, Leagues Cup, Concacaf. List the team once."),
        ("", ""),
        ("Options tab", "Scope, Option, Value. Scope is All or one sport."),
        ("enabled", "TRUE/FALSE. Turn a whole league off."),
        ("show_all_games", "TRUE/FALSE. TRUE shows the league's entire slate."),
        ("national_tv_only", "TRUE/FALSE. Keeps national broadcasts only."),
        ("standalone_only", "TRUE/FALSE. Keeps games that are the only one in their time slot"),
        ("", "-- for the NFL that is exactly TNF, SNF, MNF, the Thanksgiving"),
        ("", "trio and the December Saturday specials, and none of the"),
        ("", "regional Sunday-window games."),
        ("include_postseason", "TRUE/FALSE. Every playoff game, whatever else is set."),
        ("playoff_race", "TRUE/FALSE. Adds up to two teams to Key opponents, from"),
        ("", "live standings: your division's leader (or your nearest chaser if"),
        ("", "you lead it), and whoever holds the last playoff spot (or the first"),
        ("", "team out, if that spot is yours). Stops once ESPN marks you"),
        ("", "eliminated."),
        ("race_from_month", "Month number the race logic starts, e.g. 11 for November."),
        ("race_until_season_end", "TRUE/FALSE. End the race window at the real end of the"),
        ("", "regular season, read from ESPN, rather than at a month boundary."),
        ("", "Pairs with race_from_month: 3 + TRUE means March 1 through the"),
        ("", "last day of the regular season."),
        ("race_last_days", "Alternative: run only this many days before the season ends."),
        ("race_only_last_spot", "TRUE/FALSE. Show only the team holding the last playoff spot,"),
        ("", "skipping the division leader."),
        ("race_until_settled", "TRUE/FALSE. Stop as soon as your team has clinched a spot or"),
        ("", "been eliminated, rather than running to the end of the window."),
        ("race_min_odds", "Playoff-odds floor as a percent, e.g. 20. Key opponents stays"),
        ("", "quiet unless your team is at least this likely to make the playoffs."),
        ("", "NFL and NBA odds come from ESPN, MLB from FanGraphs. The NHL has no"),
        ("", "usable source, so a floor set there is ignored and standings"),
        ("", "position is used instead."),
        ("", "Favorites and watchlist teams are ALWAYS kept, whatever these say."),
        ("hide_finished", "TRUE/FALSE. Drop games already over."),
        ("show_odds", "TRUE/FALSE. Spread and over/under."),
        ("show_records", "TRUE/FALSE. W-L next to team names."),
        ("timezone", "IANA name, e.g. America/New_York."),
        ("", ""),
        ("How rules combine", "A game is shown if ANY rule matches -- rules never subtract."),
        ("", "So a game can qualify by being a favorite, on the watchlist, on"),
        ("", "national TV, or part of a league you show in full."),
        ("", ""),
        ("When changes land", "The sheet is re-read every 2 hours, and on every manual run."),
        ("", "Edits show up on the next run; nothing needs restarting."),
        ("", "If the sheet is unreachable the last working copy is used, and the"),
        ("", "page says so at the top rather than quietly showing an empty list."),
        ("", ""),
        ("Blank cells", "A blank Value means 'leave it alone', never 'turn it off'."),
    ]
    for label, text in lines:
        ws.append([label, text])
        if label:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    return ws


def main():
    wb = Workbook()
    teams_tab(wb)
    options_tab(wb)
    readme_tab(wb)
    wb.save(OUT)
    print("wrote %s" % OUT)


if __name__ == "__main__":
    main()
